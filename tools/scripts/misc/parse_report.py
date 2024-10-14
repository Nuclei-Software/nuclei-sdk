import argparse
import json
import os
import openpyxl.styles
import openpyxl.utils
import openpyxl.workbook
import pyexcel as pe
import openpyxl as openpyxl

# TODO:
# 1. Make the generated excel more readable and beautiful
#    - Make column widths more suitable for the table header
#    - Make excel content with more beautiful colors and fonts

def parse_runresult(runresult_file, CAREVAR="bench", BENCHTYPE="barebench"):
    runresult = dict()
    # Load json result from runresult file runresult.json
    with open(runresult_file) as f:
        runresult = json.load(f)
    parsed_bench = dict()
    olevels = ["O0", "O1", "O2", "O3", "Ofast", "Os", "Oz"]
    # parse the bench results from runresult json
    for archcfg in runresult:
        for benchtype in runresult[archcfg]:
            # TODO only check barebench results now
            if benchtype != "barebench":
                continue
            for subtype in runresult[archcfg][benchtype]:
                sheettype = subtype
                if BENCHTYPE == "toolbench":
                    matcholv = None
                    for olv in olevels:
                        if olv in archcfg:
                            matcholv = olv
                            break
                    if matcholv is not None:
                        sheettype = "%s_%s" % (subtype, matcholv)
                # collect the results of barebench
                # TODO ignore a lot meta data such as code size and compiler info
                if sheettype not in parsed_bench:
                    parsed_bench[sheettype] = dict()
                if CAREVAR == "bench":
                    # benchval is a dict like this
                    # {"DMIPS/MHz": 1.426 } or {"CoreMark/MHz": 1.212} or {"MWIPS/MHz": 0.068}
                    benchval = runresult[archcfg][benchtype][subtype]["value"]
                    try:
                        parsed_bench[sheettype][archcfg] = list(benchval.values())[0]
                    except:
                        # No value get from value dict
                        parsed_bench[sheettype][archcfg] = ""
                elif CAREVAR.startswith("size"):
                    sizeval = runresult[archcfg][benchtype][subtype]["size"]
                    whichsize = CAREVAR.lstrip("size").split("+")
                    # by default return total size
                    if len(whichsize) == 1 and whichsize[0] not in ("text", "data", "bss"):
                        parsed_bench[sheettype][archcfg] = sizeval["total"]
                    else:
                        # sum the key found in whichsize
                        # eg. size:text+data will return the sum of text+data section
                        sizesum = 0
                        for szkey in whichsize:
                            szkey = szkey.strip()
                            if szkey in ("text", "data", "bss"):
                                sizesum += sizeval[szkey]
                        parsed_bench[sheettype][archcfg] = sizesum
    # return the parsed bench result
    return parsed_bench

def parse_genreport(rptdir, BENCHTYPE="barebench", CAREVAR="bench"):
    # get the list of directory list in rptdir
    rv32_bench = dict()
    rv64_bench = dict()
    for d in os.listdir(rptdir):
        if d.startswith("."):
            continue
        # not a valid barebench directory
        if os.path.isdir(os.path.join(rptdir, d, BENCHTYPE)) == False:
            continue
        runresult_json_file = os.path.join(rptdir, d, BENCHTYPE, "runresult.json")

        if d.startswith("ux") or d.startswith("nx"):
            rv64_bench[d] = parse_runresult(runresult_json_file, BENCHTYPE=BENCHTYPE, CAREVAR=CAREVAR)
        else:
            rv32_bench[d] = parse_runresult(runresult_json_file, BENCHTYPE=BENCHTYPE, CAREVAR=CAREVAR)
    return rv32_bench, rv64_bench


def shorten_bitname(bitname):
    first_part = bitname.split("_config")[0]
    last_part = bitname.split("_")[-1][4:-4]
    shortname = ""
    # if "_best" in first_part:
    #     first_part = first_part.replace("_best", "_b")
    # if "_dual" in first_part:
    #     first_part = first_part.replace("_dual", "_DI")
    if "single" in bitname:
        shortname = "%s_SI-%s" % (first_part, last_part)
    else:
        shortname = "%s-%s" % (first_part, last_part)

    return shortname

def generate_excel_from_bench(rvbench, xlsname):
    # generate excel using rvbench
    rvsheets = dict()
    for bitname in rvbench:
        for subtype in rvbench[bitname]:
            if subtype not in rvsheets:
                rvsheets[subtype] = set()
            rvsheets[subtype] = set(rvbench[bitname][subtype].keys()).union(rvsheets[subtype])
    rvwbdict = dict()
    for subtype in rvsheets:
        rvsheets[subtype] = sorted(rvsheets[subtype])
    parsed_configs = []
    for bitname in rvbench:
        shortbitname = shorten_bitname(bitname)
        if shortbitname in parsed_configs:
            continue
        parsed_configs.append(shortbitname)
        for subtype in rvbench[bitname]:
            if subtype not in rvwbdict:
                rvwbdict[subtype] = [["config"]]
                for cfgname in rvsheets[subtype]:
                    rvwbdict[subtype].append([cfgname])
            rvwbdict[subtype][0].extend([shorten_bitname(bitname)])
            cfgindex = 1
            for cfgname in rvsheets[subtype]:
                if cfgname in rvbench[bitname][subtype]:
                    rvwbdict[subtype][cfgindex].extend([rvbench[bitname][subtype][cfgname]])
                else:
                    rvwbdict[subtype][cfgindex].extend([""])
                cfgindex += 1
    print("Save Excel as %s\n" % (xlsname))
    # book = pe.isave_book_as(bookdict=rvwbdict, dest_file_name=xlsname)
    book = pe.get_book(bookdict=rvwbdict)
    print(book)
    book.save_as(xlsname)
    book.save_as(xlsname + ".html")
    pass

def beautify_excel(excelfile):
    wb = openpyxl.load_workbook(filename=excelfile)
    tools = ["nuclei_gnu", "terapines", "nuclei_llvm"]
    for ws in wb.worksheets:
        # 设置字体样式
        bold_font_style = openpyxl.styles.Font(name="阿里巴巴普惠体 3.0 55 Regular", bold=True, size=11)
        font_stype = openpyxl.styles.Font(name="阿里巴巴普惠体 3.0 55 Regular", bold=False, size=11)
        alignment_style = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        ## 设置不同工具链配置下的背景色
        # 淡红色
        def_fill_style = openpyxl.styles.PatternFill(start_color="da9694", end_color="da9694", fill_type="solid")
        # 黄色
        gnu_fill_style = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # 浅蓝色
        llvm_fill_style = openpyxl.styles.PatternFill(start_color="00b0f0", end_color="00b0f0", fill_type="solid")
        # 浅绿色
        zcc_fill_style = openpyxl.styles.PatternFill(start_color="92d050", end_color="92d050", fill_type="solid")
        fill_styles = {
            "nuclei_gnu": gnu_fill_style,
            "nuclei_llvm": llvm_fill_style,
            "terapines": zcc_fill_style,
            "default": def_fill_style
        }
        # 设置第一行第一列的指定宽度和高度
        # 设置第一行自动换行，第二列以后的固定宽度为10
        ws.row_dimensions[1].height = 60
        ws.column_dimensions['A'].width = 50
        for colidx in range(1, ws.max_row):
            ws.column_dimensions[openpyxl.utils.get_column_letter(colidx+1)].width = 10
        cfgnames = None
        # 遍历每个sheet, 第一行，第一列加粗
        # 然后从第二列开始，将每一列中最大值加粗
        # 且每一列按照 nuclei_gnu, nuclei_llvm, terapines这样的分类找到最大值，并给最大值给背景色，背景色的设定参见 fill_styles
        for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            maxvals = dict()
            if col[0].column == 1:
                cfgnames = [cell.value for cell in col]
            else:
                colvals = [cell.value for cell in col]
                for (cfg, value) in zip(cfgnames, colvals):
                    if cfg == "config":
                        continue
                    newval = value if value else 0
                    matchedtool = "default"
                    for tool in tools:
                        if tool in cfg:
                            matchedtool = tool
                            break
                    if matchedtool not in maxvals:
                        maxvals[matchedtool] = newval
                    else:
                        maxvals[matchedtool] = max(maxvals[matchedtool], newval)
            maxval = 0
            if len(maxvals) > 0:
                maxval = max(list(maxvals.values()))
            for (cfg, cell) in zip(cfgnames, col):
                cell.alignment = alignment_style
                if cell.row == 1 or cell.column == 1:
                    cell.font = bold_font_style
                else:
                    cell.font = font_stype
                if cell.column > 1 and cell.row > 1:
                    matchedtool = "default"
                    for tool in tools:
                        if tool in cfg:
                            matchedtool = tool
                            break
                    if cell.value and maxvals[matchedtool] == cell.value:
                        cell.fill = fill_styles[matchedtool]
                        if maxval == cell.value: # 让最大值加粗
                            cell.font = bold_font_style
    wb.save(excelfile.replace(".xlsx", "_styled.xlsx"))
    pass
def generate_bench_excel(rptdir, BENCHTYPE="barebench", CAREVAR="bench"):
    rv32_bench, rv64_bench = parse_genreport(rptdir, BENCHTYPE=BENCHTYPE, CAREVAR=CAREVAR)
    # generate excel using rv32_bench
    generate_excel_from_bench(rv32_bench, os.path.join(rptdir, "rv32_%s.xlsx" % (BENCHTYPE)))
    beautify_excel(os.path.join(rptdir, "rv32_%s.xlsx" % (BENCHTYPE)))
    generate_excel_from_bench(rv64_bench, os.path.join(rptdir, "rv64_%s.xlsx" % (BENCHTYPE)))
    beautify_excel(os.path.join(rptdir, "rv64_%s.xlsx" % (BENCHTYPE)))
    pass

# 使用argpaser 解析参数，-d 表示待解析的目录
parser = argparse.ArgumentParser(description='Parse the nsdk_cli run reports')
parser.add_argument('-d', '--dir', help='directory to parse')

args = parser.parse_args()

reportdir = args.dir.strip()
for benchtype in ["barebench", "toolbench"]:
    generate_bench_excel(reportdir, benchtype)
